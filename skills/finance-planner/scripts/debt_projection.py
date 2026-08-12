#!/usr/bin/env python3
"""
Debt-free projection — cascade simulator (Money Mentor / finance-planner skill).

Reads active debts from the FinPlan Debt Tracker sheet (or a CSV), then simulates
month-by-month: interest accrues (balance x rate / 12), minimums are paid, the
snowball pool attacks the highest-priority living debt, and every kill rolls its
freed minimum into the pool from the next month.

Usage (run from the Finance folder root):
  python3 money-mentor-plugin/skills/finance-planner/scripts/debt_projection.py \
      --xlsx spreadsheets/FinPlan_v2.xlsx --pool 13971 --reserve 750 \
      --start 2026-06 --inject 2026-08:19300 --label "with tax refund"

Key arguments:
  --pool     REQUIRED. Net Snowball Power (realistic) from household-data.md
             Current State at run date. The cascade adds freed minimums on top
             as debts die DURING the simulation — do not pre-add them.
  --reserve  Monthly standing diversion (e.g. vehicle reserve). Default 0.
  --inject   One-off boost "YYYY-MM:amount" (tax refund, bonus). Repeatable.
  --divert   "YYYY-MM:permonth:target" — divert per-month into a named goal
             (e.g. move fund) until target reached (e.g. 2027-02:8000:65000).
  --include-vehicles  Also cascade into vehicle debts after consumer debts.
             (Indicative only — a planned vehicle sale overrides this; see
             vehicle-strategy notes in the household folder.)
  --csv      Alternative input: CSV with header name,balance,rate,min,priority
             (rate as annual decimal, e.g. 0.2725).

Output: markdown — kill-date table, interest totals, minimums-only contrast,
interest-burn trajectory. Paste/refresh into archive/YYYY/debt_projection_*.md.
"""
import argparse, csv, sys
from datetime import date

VEHICLE_HINTS = ("jeep", "kia", "wesbank", "mfc", "vehicle", "car ")

def month_add(d, n):
    m = d.month - 1 + n
    return date(d.year + m // 12, m % 12 + 1, 1)

def fmt(d):
    return d.strftime("%b %Y")

def parse_month(s):
    y, m = s.split("-")
    return date(int(y), int(m), 1)

def load_xlsx(path):
    import openpyxl
    ws = openpyxl.load_workbook(path, data_only=True)["Debt Tracker"]
    debts = []
    for r in range(3, 13):
        name = ws.cell(row=r, column=2).value
        bal = ws.cell(row=r, column=4).value or 0
        rate = ws.cell(row=r, column=5).value or 0
        minp = ws.cell(row=r, column=6).value or 0
        prio = ws.cell(row=r, column=7).value or r
        if name and float(bal) > 0.01:
            debts.append(dict(name=str(name), bal=float(bal), rate=float(rate),
                              min=float(minp), prio=float(prio)))
    return debts

def load_csv(path):
    debts = []
    with open(path) as f:
        for row in csv.DictReader(f):
            if float(row["balance"]) > 0.01:
                debts.append(dict(name=row["name"], bal=float(row["balance"]),
                                  rate=float(row["rate"]), min=float(row["min"]),
                                  prio=float(row.get("priority", 99))))
    return debts

def is_vehicle(d):
    return any(h in d["name"].lower() for h in VEHICLE_HINTS)

def simulate(debts, pool, reserve, start, injects, divert, max_months=240):
    debts = sorted([dict(d) for d in debts], key=lambda d: d["prio"])
    cur, months = start, 0
    kills, interest_total, burn_points = [], 0.0, []
    div_fund = 0.0
    while any(d["bal"] > 0.01 for d in debts) and months < max_months:
        extra = max(0.0, pool - reserve) + injects.get(cur, 0.0)
        if divert and cur >= divert["from"] and div_fund < divert["target"]:
            take = min(divert["per"], divert["target"] - div_fund, extra)
            div_fund += take
            extra -= take
        burn = 0.0
        for d in debts:
            if d["bal"] <= 0.01:
                continue
            i = d["bal"] * d["rate"] / 12
            burn += i
            interest_total += i
            d["bal"] += i
            d["bal"] -= min(d["bal"], d["min"])
            if d["bal"] > 0.01 and extra > 0:
                # cascade with same-month overflow: walking in priority order,
                # leftover extra flows to the next living debt this month
                pay = min(d["bal"], extra)
                d["bal"] -= pay
                extra -= pay
            if d["bal"] <= 0.01:
                kills.append((d["name"], cur, d["min"]))
                pool += d["min"]          # freed minimum joins pool next month
        if months % 3 == 0 or all(d["bal"] <= 0.01 for d in debts):
            burn_points.append((cur, burn))
        cur = month_add(cur, 1)
        months += 1
    done = kills[-1][1] if kills and all(d["bal"] <= 0.01 for d in debts) else None
    return kills, interest_total, done, div_fund, burn_points

def minimums_only(debts, start, max_months=600):
    debts = [dict(d) for d in debts]
    cur, months, interest_total = start, 0, 0.0
    while any(d["bal"] > 0.01 for d in debts) and months < max_months:
        for d in debts:
            if d["bal"] <= 0.01:
                continue
            i = d["bal"] * d["rate"] / 12
            interest_total += i
            d["bal"] += i
            d["bal"] -= min(d["bal"], d["min"])
        cur = month_add(cur, 1)
        months += 1
    return cur, interest_total, months < max_months

def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx")
    ap.add_argument("--csv")
    ap.add_argument("--pool", type=float, required=True)
    ap.add_argument("--reserve", type=float, default=0.0)
    ap.add_argument("--start", default=None, help="YYYY-MM (default: next month)")
    ap.add_argument("--inject", action="append", default=[], help="YYYY-MM:amount")
    ap.add_argument("--divert", default=None, help="YYYY-MM:permonth:target")
    ap.add_argument("--include-vehicles", action="store_true")
    ap.add_argument("--label", default="projection")
    a = ap.parse_args()

    if not (a.xlsx or a.csv):
        sys.exit("Provide --xlsx or --csv")
    debts = load_xlsx(a.xlsx) if a.xlsx else load_csv(a.csv)
    consumer = [d for d in debts if not is_vehicle(d)]
    vehicles = [d for d in debts if is_vehicle(d)]
    sim_debts = consumer + (vehicles if a.include_vehicles else [])
    if not sim_debts:
        sys.exit("No active debts found.")

    today = date.today()
    start = parse_month(a.start) if a.start else month_add(date(today.year, today.month, 1), 1)
    injects = {}
    for s in a.inject:
        m, amt = s.split(":")
        injects[parse_month(m)] = injects.get(parse_month(m), 0.0) + float(amt)
    divert = None
    if a.divert:
        m, per, tgt = a.divert.split(":")
        divert = dict(from_=None, per=float(per), target=float(tgt))
        divert["from"] = parse_month(m)

    kills, intr, done, fund, burns = simulate(sim_debts, a.pool, a.reserve, start, injects, divert)
    mo_end, mo_intr, mo_ok = minimums_only(sim_debts, start)

    print(f"# Debt projection — {a.label} (run {today.isoformat()})")
    print(f"\nInputs: pool R{a.pool:,.0f}/mo · reserve R{a.reserve:,.0f}/mo · start {fmt(start)}"
          + (f" · injections {', '.join(f'{fmt(k)} R{v:,.0f}' for k, v in injects.items())}" if injects else "")
          + (f" · divert R{divert['per']:,.0f}/mo from {fmt(divert['from'])} to R{divert['target']:,.0f}" if divert else "")
          + (" · vehicles INCLUDED" if a.include_vehicles else " · consumer debts only"))
    print("\n| Debt | Dead | Frees /mo |\n|---|---|---|")
    for name, when, minp in kills:
        print(f"| {name} | **{fmt(when)}** | R{minp:,.2f} |")
    if done:
        print(f"\n**ALL SIMULATED DEBT DEAD: {fmt(done)}** (±1–2 months for slip months)")
    else:
        print("\n⚠ Did not complete within horizon — check inputs.")
    if divert:
        print(f"Diverted goal fund at end: R{fund:,.0f} of R{divert['target']:,.0f}")
    print(f"\nInterest paid in plan: **R{intr:,.0f}**")
    if mo_ok:
        print(f"Minimums-only contrast: dead {fmt(mo_end)}, interest R{mo_intr:,.0f} "
              f"→ strategy saves **R{mo_intr - intr:,.0f}** and finishes "
              f"{(mo_end.year - done.year) * 12 + (mo_end.month - done.month) if done else '—'} months sooner.")
    else:
        print("Minimums-only contrast: NEVER finishes within 50 years (minimums don't cover interest on some debt).")
    print("\nInterest-burn trajectory (quarterly):")
    print(" → ".join(f"{fmt(m)}: R{b:,.0f}" for m, b in burns))
    print("\n*Assumes the pool shows up every month; slip months shift dates ~1:1. "
          "Re-run at every month close with --pool set to the current realistic snowball "
          "from household-data.md. Vehicle phase is indicative — planned sales override.*")

if __name__ == "__main__":
    main()
